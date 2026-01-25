#!/usr/bin/env python3
"""
Generate bulk import file for students compatible with the bulk import API.
Creates 30 students: 10 in Morning section, 10 in Evening section, 10 in Computer engineering class only.

Format: name,email,phone,password,instituteId,classId,sectionId,courseId
"""

import csv
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not found. Excel generation will be skipped.")
    print("Install with: pip install openpyxl")

# IDs from the Test University structure (CORRECTED - note the 'S' after 'KF')
INSTITUTE_ID = "01KFSRV0T1D39E20016FB874CB"
CLASS_ID = "01KFSZ83HQ0A66BDD01A0C3DB4"

SECTION_MORNING_ID = "01KFSZ83J9CE865416EBF08497"
SECTION_EVENING_ID = "01KFSZ83JE92BAFC69C1D130B3"

# Configuration
STUDENTS_PER_GROUP = 10
EMAIL_DOMAIN = "testuni.com"
COURSE_ID = ""  # Leave empty to use class/section association only


def generate_students_data():
    """Generate student data for bulk import"""
    students = []
    student_counter = 1
    
    # Group 1: Morning section students (1-10)
    for i in range(STUDENTS_PER_GROUP):
        student = {
            "name": f"Student{student_counter} Morning",
            "email": f"student{student_counter}@{EMAIL_DOMAIN}",
            "phone": f"+1555{student_counter:04d}",
            "password": "",  # Will be auto-generated
            "instituteId": INSTITUTE_ID,
            "classId": CLASS_ID,
            "sectionId": SECTION_MORNING_ID,
            "courseId": COURSE_ID
        }
        students.append(student)
        student_counter += 1
    
    # Group 2: Evening section students (11-20)
    for i in range(STUDENTS_PER_GROUP):
        student = {
            "name": f"Student{student_counter} Evening",
            "email": f"student{student_counter}@{EMAIL_DOMAIN}",
            "phone": f"+1555{student_counter:04d}",
            "password": "",  # Will be auto-generated
            "instituteId": INSTITUTE_ID,
            "classId": CLASS_ID,
            "sectionId": SECTION_EVENING_ID,
            "courseId": COURSE_ID
        }
        students.append(student)
        student_counter += 1
    
    # Group 3: Class-level only students (21-30) - no sectionId
    for i in range(STUDENTS_PER_GROUP):
        student = {
            "name": f"Student{student_counter} ClassOnly",
            "email": f"student{student_counter}@{EMAIL_DOMAIN}",
            "phone": f"+1555{student_counter:04d}",
            "password": "",  # Will be auto-generated
            "instituteId": INSTITUTE_ID,
            "classId": CLASS_ID,
            "sectionId": "",  # Empty for class-level only
            "courseId": COURSE_ID
        }
        students.append(student)
        student_counter += 1
    
    return students


def generate_csv(filename="bulk_import_students.csv"):
    """Generate CSV file for bulk import"""
    students = generate_students_data()
    
    # CSV column headers (must match bulk import format)
    headers = ["name", "email", "phone", "password", "instituteId", "classId", "sectionId", "courseId"]
    
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        writer.writeheader()
        writer.writerows(students)
    
    print(f"✅ CSV file created: {filename}")
    print(f"📊 Total students: {len(students)}")
    print(f"   - Morning section: {STUDENTS_PER_GROUP}")
    print(f"   - Evening section: {STUDENTS_PER_GROUP}")
    print(f"   - Class level only: {STUDENTS_PER_GROUP}")
    return filename


def generate_excel(filename="bulk_import_students.xlsx"):
    """Generate Excel file for bulk import"""
    if not OPENPYXL_AVAILABLE:
        print("❌ Cannot generate Excel: openpyxl not installed")
        return None
    
    students = generate_students_data()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students Import"
    
    # Headers
    headers = ["name", "email", "phone", "password", "instituteId", "classId", "sectionId", "courseId"]
    
    # Style for header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style for different sections
    morning_fill = PatternFill(start_color="E7E6FF", end_color="E7E6FF", fill_type="solid")
    evening_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    class_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Write student data
    for row_idx, student in enumerate(students, start=2):
        # Determine fill color based on student group
        if row_idx <= STUDENTS_PER_GROUP + 1:
            fill_color = morning_fill
        elif row_idx <= (STUDENTS_PER_GROUP * 2) + 1:
            fill_color = evening_fill
        else:
            fill_color = class_fill
        
        for col_idx, header in enumerate(headers, start=1):
            value = student[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill_color
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add instructions sheet
    ws_info = wb.create_sheet("Import Instructions", 0)
    
    instructions = [
        ["Bulk Student Import File", ""],
        ["", ""],
        ["Instructions:", ""],
        ["1. Fill in the 'Students Import' sheet with student data", ""],
        ["2. Required fields: name, email, instituteId", ""],
        ["3. Password: Leave empty to auto-generate", ""],
        ["4. Section: Leave sectionId empty for class-level only association", ""],
        ["5. Upload this file to the Bulk Import page in the admin dashboard", ""],
        ["", ""],
        ["Field Descriptions:", ""],
        ["name", "Full name of the student (required)"],
        ["email", "Email address - must be unique (required)"],
        ["phone", "Phone number (optional)"],
        ["password", "Password - leave empty to auto-generate (optional)"],
        ["instituteId", "ID of the institute (required)"],
        ["classId", "ID of the class (optional but recommended)"],
        ["sectionId", "ID of the section - leave empty for class-level only (optional)"],
        ["courseId", "ID of the course to enroll in (optional)"],
        ["", ""],
        ["Test University IDs:", ""],
        ["Institute ID:", INSTITUTE_ID],
        ["Class ID (Computer engineering):", CLASS_ID],
        ["Morning Section ID:", SECTION_MORNING_ID],
        ["Evening Section ID:", SECTION_EVENING_ID],
        ["", ""],
        ["Student Distribution:", ""],
        ["Rows 2-11:", "Morning section (10 students)"],
        ["Rows 12-21:", "Evening section (10 students)"],
        ["Rows 22-31:", "Class level only, no section (10 students)"],
        ["", ""],
        ["Import Options (in UI):", ""],
        ["✓ Auto-generate passwords", "Recommended - enabled by default"],
        ["✓ Update existing students", "Optional - updates if email exists"],
        ["✓ Auto-enroll in courses", "Optional - only if courseId is provided"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1 and value and not value.startswith(" "):
                cell.font = Font(bold=True)
    
    ws_info.column_dimensions['A'].width = 40
    ws_info.column_dimensions['B'].width = 50
    
    # Save workbook
    wb.save(filename)
    print(f"✅ Excel file created: {filename}")
    print(f"📊 Total students: {len(students)}")
    print(f"   - Morning section: {STUDENTS_PER_GROUP}")
    print(f"   - Evening section: {STUDENTS_PER_GROUP}")
    print(f"   - Class level only: {STUDENTS_PER_GROUP}")
    return filename


def print_summary():
    """Print summary of what will be generated"""
    print("\n" + "="*60)
    print("BULK IMPORT FILE GENERATOR")
    print("="*60)
    print(f"\nInstitute ID: {INSTITUTE_ID}")
    print(f"Class ID:     {CLASS_ID}")
    print(f"Morning Section ID: {SECTION_MORNING_ID}")
    print(f"Evening Section ID: {SECTION_EVENING_ID}")
    print(f"\nTotal Students: {STUDENTS_PER_GROUP * 3}")
    print(f"  - Morning section:  {STUDENTS_PER_GROUP} (student1@{EMAIL_DOMAIN} - student{STUDENTS_PER_GROUP}@{EMAIL_DOMAIN})")
    print(f"  - Evening section:  {STUDENTS_PER_GROUP} (student{STUDENTS_PER_GROUP + 1}@{EMAIL_DOMAIN} - student{STUDENTS_PER_GROUP * 2}@{EMAIL_DOMAIN})")
    print(f"  - Class level only: {STUDENTS_PER_GROUP} (student{STUDENTS_PER_GROUP * 2 + 1}@{EMAIL_DOMAIN} - student{STUDENTS_PER_GROUP * 3}@{EMAIL_DOMAIN})")
    print("\nFile format: CSV compatible with bulk import API")
    print("Passwords: Will be auto-generated during import (leave empty)")
    print("="*60 + "\n")


if __name__ == "__main__":
    print_summary()
    
    # Generate both CSV and Excel
    csv_file = generate_csv()
    print()
    
    if OPENPYXL_AVAILABLE:
        excel_file = generate_excel()
        print()
        print("📝 Files ready for bulk import!")
        print(f"   - CSV:   {csv_file}")
        print(f"   - Excel: {excel_file}")
    else:
        print("📝 CSV file ready for bulk import!")
        print(f"   - CSV: {csv_file}")
        print("\n💡 Tip: Install openpyxl to generate Excel files too:")
        print("   pip install openpyxl")
    
    print("\n🚀 Next steps:")
    print("1. Open the admin dashboard")
    print("2. Navigate to Students → Bulk Import")
    print("3. Upload the generated file")
    print("4. Enable 'Auto-generate passwords'")
    print("5. Optionally enable 'Update existing students'")
    print("6. Click 'Import Students'")
