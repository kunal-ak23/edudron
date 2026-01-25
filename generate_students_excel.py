#!/usr/bin/env python3
"""
Script to generate test students Excel file with 3 groups:
- 10 students in Morning section
- 10 students in Evening section
- 10 students directly associated with Computer engineering class (no section)

Email format: student<no>@testuni.com
"""

import uuid
from datetime import datetime, timezone
import json

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not found. Installing...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# IDs from the Test University structure
INSTITUTE_ID = "01KF5RV0T1D39E20016FB874CB"
INSTITUTE_NAME = "Test Uni Inst"
CLASS_ID = "01KF5Z83HQ0A66BDD01A0C3DB4"
CLASS_NAME = "Computer engineering"
CLASS_CODE = "COM2024"

SECTIONS = [
    {"id": "01KF5Z83J9CE865416EBF08497", "name": "Morning"},
    {"id": "01KF5Z83JE92BAFC69C1D130B3", "name": "Evening"}
]

# Configuration
STUDENTS_PER_GROUP = 10
EMAIL_DOMAIN = "testuni.com"
CLIENT_ID = "YOUR_CLIENT_ID_HERE"  # Replace with actual tenant/client UUID
COURSE_ID = "YOUR_COURSE_ID_HERE"  # Replace with actual course UUID if needed


def generate_ulid():
    """Generate a ULID-style ID (simplified version)"""
    timestamp = int(datetime.now(timezone.utc).timestamp() * 1000)
    random_part = uuid.uuid4().hex[:16].upper()
    return f"{timestamp:013X}{random_part}"


def generate_students_data():
    """Generate student data for all three groups"""
    
    student_counter = 1
    all_students = []
    
    # Group 1: Morning section students (10)
    for i in range(STUDENTS_PER_GROUP):
        student_id = generate_ulid()
        email = f"student{student_counter}@{EMAIL_DOMAIN}"
        
        student_data = {
            "student_id": student_id,
            "email": email,
            "first_name": f"Student{student_counter}",
            "last_name": "Morning",
            "phone": f"+1555{student_counter:04d}",
            "institute_id": INSTITUTE_ID,
            "institute_name": INSTITUTE_NAME,
            "class_id": CLASS_ID,
            "class_name": CLASS_NAME,
            "class_code": CLASS_CODE,
            "section_id": SECTIONS[0]["id"],
            "section_name": SECTIONS[0]["name"],
            "association_type": "Section Level",
            "enrollment_id": generate_ulid(),
            "is_active": True
        }
        
        all_students.append(student_data)
        student_counter += 1
    
    # Group 2: Evening section students (10)
    for i in range(STUDENTS_PER_GROUP):
        student_id = generate_ulid()
        email = f"student{student_counter}@{EMAIL_DOMAIN}"
        
        student_data = {
            "student_id": student_id,
            "email": email,
            "first_name": f"Student{student_counter}",
            "last_name": "Evening",
            "phone": f"+1555{student_counter:04d}",
            "institute_id": INSTITUTE_ID,
            "institute_name": INSTITUTE_NAME,
            "class_id": CLASS_ID,
            "class_name": CLASS_NAME,
            "class_code": CLASS_CODE,
            "section_id": SECTIONS[1]["id"],
            "section_name": SECTIONS[1]["name"],
            "association_type": "Section Level",
            "enrollment_id": generate_ulid(),
            "is_active": True
        }
        
        all_students.append(student_data)
        student_counter += 1
    
    # Group 3: Class-level only students (no section) (10)
    for i in range(STUDENTS_PER_GROUP):
        student_id = generate_ulid()
        email = f"student{student_counter}@{EMAIL_DOMAIN}"
        
        student_data = {
            "student_id": student_id,
            "email": email,
            "first_name": f"Student{student_counter}",
            "last_name": "ClassOnly",
            "phone": f"+1555{student_counter:04d}",
            "institute_id": INSTITUTE_ID,
            "institute_name": INSTITUTE_NAME,
            "class_id": CLASS_ID,
            "class_name": CLASS_NAME,
            "class_code": CLASS_CODE,
            "section_id": None,
            "section_name": "No Section (Class Level)",
            "association_type": "Class Level Only",
            "enrollment_id": generate_ulid(),
            "is_active": True
        }
        
        all_students.append(student_data)
        student_counter += 1
    
    return all_students


def create_excel_file(filename="test_students.xlsx"):
    """Create an Excel file with student data"""
    
    students = generate_students_data()
    
    # Create workbook and worksheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create worksheets
    ws_summary = wb.create_sheet("Summary", 0)
    ws_all = wb.create_sheet("All Students", 1)
    ws_morning = wb.create_sheet("Morning Section", 2)
    ws_evening = wb.create_sheet("Evening Section", 3)
    ws_class_only = wb.create_sheet("Class Level Only", 4)
    ws_sql = wb.create_sheet("SQL Inserts", 5)
    
    # Define headers
    headers = [
        "Student ID", "Email", "First Name", "Last Name", "Phone",
        "Institute ID", "Institute Name", "Class ID", "Class Name", "Class Code",
        "Section ID", "Section Name", "Association Type", "Enrollment ID", "Active"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    morning_fill = PatternFill(start_color="E7E6FF", end_color="E7E6FF", fill_type="solid")
    evening_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    class_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # ===== SUMMARY SHEET =====
    ws_summary.title = "Summary"
    
    # Add title (merge cells after setting value)
    ws_summary['A1'] = "Test University Student Generation Summary"
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    summary_data = [
        ["Generated At:", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")],
        ["", ""],
        ["Institute:", INSTITUTE_NAME],
        ["Institute ID:", INSTITUTE_ID],
        ["", ""],
        ["Class:", CLASS_NAME],
        ["Class Code:", CLASS_CODE],
        ["Class ID:", CLASS_ID],
        ["", ""],
        ["Morning Section ID:", SECTIONS[0]["id"]],
        ["Evening Section ID:", SECTIONS[1]["id"]],
        ["", ""],
        ["Total Students:", len(students)],
        ["Morning Section Students:", STUDENTS_PER_GROUP],
        ["Evening Section Students:", STUDENTS_PER_GROUP],
        ["Class Level Only Students:", STUDENTS_PER_GROUP],
        ["", ""],
        ["Email Format:", f"student<no>@{EMAIL_DOMAIN}"],
        ["Email Range:", f"student1@{EMAIL_DOMAIN} to student{len(students)}@{EMAIL_DOMAIN}"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1 and value:
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 50
    
    # ===== ALL STUDENTS SHEET =====
    def populate_sheet(ws, students_subset, title_color=None):
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for row_idx, student in enumerate(students_subset, start=2):
            row_data = [
                student["student_id"],
                student["email"],
                student["first_name"],
                student["last_name"],
                student["phone"],
                student["institute_id"],
                student["institute_name"],
                student["class_id"],
                student["class_name"],
                student["class_code"],
                student["section_id"] if student["section_id"] else "NULL",
                student["section_name"],
                student["association_type"],
                student["enrollment_id"],
                "TRUE" if student["is_active"] else "FALSE"
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if title_color:
                    cell.fill = title_color
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 25
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    populate_sheet(ws_all, students)
    populate_sheet(ws_morning, [s for s in students if s["section_name"] == "Morning"], morning_fill)
    populate_sheet(ws_evening, [s for s in students if s["section_name"] == "Evening"], evening_fill)
    populate_sheet(ws_class_only, [s for s in students if s["section_id"] is None], class_fill)
    
    # ===== SQL INSERTS SHEET =====
    sql_lines = [
        "-- =====================================================",
        "-- SQL INSERT Statements for Test Students",
        f"-- Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
        "-- =====================================================",
        "",
        "-- IMPORTANT: Replace 'YOUR_CLIENT_ID_HERE' with your actual tenant UUID",
        "-- IMPORTANT: Replace 'YOUR_COURSE_ID_HERE' with your actual course UUID",
        "",
        "-- =====================================================",
        "-- INSERT STUDENTS",
        "-- =====================================================",
        "",
        "INSERT INTO student.students (",
        "    id, client_id, email, first_name, last_name, phone,",
        "    is_active, created_at, updated_at",
        ") VALUES"
    ]
    
    for idx, student in enumerate(students):
        comma = "," if idx < len(students) - 1 else ";"
        sql_lines.append(
            f"    ('{student['student_id']}', '{CLIENT_ID}', '{student['email']}', "
            f"'{student['first_name']}', '{student['last_name']}', '{student['phone']}', "
            f"true, NOW(), NOW()){comma}"
        )
    
    sql_lines.extend([
        "",
        "-- =====================================================",
        "-- INSERT ENROLLMENTS",
        "-- =====================================================",
        "",
        "INSERT INTO student.enrollments (",
        "    id, client_id, student_id, institute_id, class_id, batch_id, course_id,",
        "    enrolled_at, created_at, updated_at",
        ") VALUES"
    ])
    
    for idx, student in enumerate(students):
        comma = "," if idx < len(students) - 1 else ";"
        section_id = f"'{student['section_id']}'" if student['section_id'] else "NULL"
        sql_lines.append(
            f"    ('{student['enrollment_id']}', '{CLIENT_ID}', '{student['student_id']}', "
            f"'{student['institute_id']}', '{student['class_id']}', {section_id}, "
            f"'{COURSE_ID}', NOW(), NOW(), NOW()){comma}"
        )
    
    # Write SQL to sheet
    for row_idx, line in enumerate(sql_lines, start=1):
        ws_sql.cell(row=row_idx, column=1, value=line)
    
    ws_sql.column_dimensions['A'].width = 120
    
    # Save workbook
    wb.save(filename)
    print(f"✅ Excel file created: {filename}")
    print(f"📊 Total students: {len(students)}")
    print(f"   - Morning section: {len([s for s in students if s['section_name'] == 'Morning'])}")
    print(f"   - Evening section: {len([s for s in students if s['section_name'] == 'Evening'])}")
    print(f"   - Class level only: {len([s for s in students if s['section_id'] is None])}")
    
    return filename


if __name__ == "__main__":
    import sys
    
    filename = "test_students.xlsx"
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    
    create_excel_file(filename)
    print(f"\n📝 Open {filename} to view the generated students!")
